"""HRM export: per-punch rows, 序號, unpadded dates, fixed APP打卡 label."""
import datetime
import io

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attendance_log import AttendanceLog, WorkMode
from app.models.employee import Employee, Role
from app.services import reporting_service
from app.utils.password import hash_password


async def _seed_employee(
    db_session: AsyncSession,
    emp_id: str,
    role: Role | str = Role.EMPLOYEE,
) -> Employee:
    role_enum = role if isinstance(role, Role) else Role(role)
    emp = Employee(
        emp_id=emp_id,
        name=f"User {emp_id}",
        department="Engineering",
        role=role_enum,
        hashed_password=hash_password("pass1234"),
        shift_start_time=datetime.time(9, 0),
        shift_end_time=datetime.time(18, 0),
    )
    db_session.add(emp)
    await db_session.commit()
    await db_session.refresh(emp)
    return emp


def _make_log(emp_id: str, d: datetime.date, t: datetime.time) -> AttendanceLog:
    return AttendanceLog(
        emp_id=emp_id,
        timestamp=datetime.datetime.combine(d, t),
        latitude=25.0,
        longitude=121.5,
        accuracy=10.0,
        ip_address="127.0.0.1",
        work_mode=WorkMode.OFFICE,
    )


@pytest.mark.asyncio
async def test_hrm_export_two_rows_per_full_day(db_session):
    # Real attendance_logs (not just a summary row) so NORMAL status survives
    # the export pipeline's recomputation via generate_all_summaries.
    await _seed_employee(db_session, "E960")
    d = datetime.date(2026, 7, 24)
    db_session.add_all([
        _make_log("E960", d, datetime.time(8, 30)),
        _make_log("E960", d, datetime.time(17, 30)),
    ])
    await db_session.commit()

    content = await reporting_service.export_attendance_hrm(
        db_session, start_date=d, end_date=d, submission_filter="all")
    ws = load_workbook(io.BytesIO(content)).active

    header = [c.value for c in ws[1]]
    assert header == ["序號", "工號", "姓名", "考勤機ID", "刷卡日期", "刷卡時間",
                      "補刷卡假勤類型原因"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert rows == [
        [1, "E960", "User E960", None, "2026/7/24", "08:30", "APP打卡"],
        [2, "E960", "User E960", None, "2026/7/24", "17:30", "APP打卡"],
    ]


@pytest.mark.asyncio
async def test_hrm_export_single_punch_one_row_and_absent_none(db_session):
    await _seed_employee(db_session, "E961")
    d = datetime.date(2026, 7, 24)
    # single punch: first == last → one row only
    db_session.add(_make_log("E961", d, datetime.time(8, 30)))
    await db_session.commit()
    content = await reporting_service.export_attendance_hrm(
        db_session, start_date=d, end_date=d, submission_filter="all")
    ws = load_workbook(io.BytesIO(content)).active
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert len(rows) == 1 and rows[0][5] == "08:30"


@pytest.mark.asyncio
async def test_hrm_export_excludes_seed_accounts(db_session):
    await _seed_employee(db_session, "HR01", role=Role.HR)
    d = datetime.date(2026, 7, 24)
    db_session.add_all([
        _make_log("HR01", d, datetime.time(8, 30)),
        _make_log("HR01", d, datetime.time(17, 30)),
    ])
    await db_session.commit()
    content = await reporting_service.export_attendance_hrm(
        db_session, start_date=d, end_date=d, submission_filter="all")
    ws = load_workbook(io.BytesIO(content)).active
    assert list(ws.iter_rows(min_row=2)) == []
