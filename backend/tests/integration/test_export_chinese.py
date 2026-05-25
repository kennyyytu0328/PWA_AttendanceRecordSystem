"""Tests for Chinese-ized export with new columns + submission filter."""
import csv
import datetime
import io

from app.models.daily_attendance_summary import AttendanceStatus
from app.models.employee import Employee, Role
from app.repositories import (
    monthly_submission_repository,
    summary_repository,
)
from app.services import reporting_service


async def _seed_employee_with_shift(
    session,
    emp_id: str,
    *,
    name: str = "Test User",
    department: str = "Eng",
    shift_start_time: str = "09:00",
    shift_end_time: str = "18:00",
) -> Employee:
    """Insert an employee with the given shift window (HH:MM strings)."""
    start_h, start_m = (int(p) for p in shift_start_time.split(":"))
    end_h, end_m = (int(p) for p in shift_end_time.split(":"))
    emp = Employee(
        emp_id=emp_id,
        name=name,
        department=department,
        role=Role.EMPLOYEE,
        hashed_password="x",
        shift_start_time=datetime.time(start_h, start_m),
        shift_end_time=datetime.time(end_h, end_m),
    )
    session.add(emp)
    await session.commit()
    await session.refresh(emp)
    return emp


async def test_csv_export_uses_chinese_headers(db_session):
    await _seed_employee_with_shift(
        db_session, emp_id="E040", shift_start_time="09:00", shift_end_time="18:00"
    )
    await summary_repository.upsert_summary(
        db_session,
        emp_id="E040",
        date=datetime.date(2026, 5, 14),
        first_clock_in=datetime.datetime(2026, 5, 14, 9, 0),
        last_clock_out=datetime.datetime(2026, 5, 14, 18, 0),
        status=AttendanceStatus.NORMAL,
    )
    await monthly_submission_repository.upsert(
        db_session, emp_id="E040", year=2026, month=5
    )

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 14),
        end_date=datetime.date(2026, 5, 14),
        format="csv",
    )
    reader = csv.reader(io.StringIO(csv_text))
    header = next(reader)
    assert header == [
        "員工編號", "姓名", "部門", "日期", "星期",
        "班別時間", "上班時間", "下班時間",
        "狀態", "備註", "加班時數", "遲到理由", "送單狀態",
    ]


async def test_csv_export_translates_status_values(db_session):
    await _seed_employee_with_shift(
        db_session, emp_id="E041", shift_start_time="09:00", shift_end_time="18:00"
    )
    await summary_repository.upsert_summary(
        db_session,
        emp_id="E041",
        date=datetime.date(2026, 5, 14),
        first_clock_in=None,
        last_clock_out=None,
        status=AttendanceStatus.LEAVE,
        leave_type="特休",
        remark="上午",
    )
    await monthly_submission_repository.upsert(
        db_session, emp_id="E041", year=2026, month=5
    )

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 14),
        end_date=datetime.date(2026, 5, 14),
        format="csv",
    )
    rows = list(csv.reader(io.StringIO(csv_text)))
    data_row = rows[1]
    assert data_row[0] == "E041"        # 員工編號
    assert data_row[4] == "星期四"        # 星期 (2026-05-14 is a Thursday)
    assert data_row[8] == "請假"          # 狀態
    assert data_row[9] == "特休·上午"     # 備註
    assert data_row[10] == ""            # 加班時數 — none
    assert data_row[12] == "已送單"       # 送單狀態


async def test_csv_export_rest_day_work_annotated_in_remark(db_session):
    """A non-makeup Saturday with overtime work is annotated 休息日加班 in 備註,
    shows 星期六 in the new weekday column, and stays 正常 in 狀態."""
    import decimal

    from app.models.attendance_log import AttendanceLog, WorkMode
    from app.repositories import system_config_repository

    await _seed_employee_with_shift(
        db_session, emp_id="E_SAT", shift_start_time="09:00", shift_end_time="18:00"
    )
    # 2026-05-09 is a Saturday → REST_DAY.
    db_session.add_all(
        [
            AttendanceLog(
                emp_id="E_SAT",
                timestamp=datetime.datetime(2026, 5, 9, 10, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
            AttendanceLog(
                emp_id="E_SAT",
                timestamp=datetime.datetime(2026, 5, 9, 12, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
        ]
    )
    await db_session.commit()
    await summary_repository.upsert_summary(
        db_session,
        emp_id="E_SAT",
        date=datetime.date(2026, 5, 9),
        first_clock_in=datetime.datetime(2026, 5, 9, 10, 0),
        last_clock_out=datetime.datetime(2026, 5, 9, 12, 0),
        status=AttendanceStatus.NORMAL,
        overtime_hours=decimal.Decimal("3"),
    )
    await monthly_submission_repository.upsert(
        db_session, emp_id="E_SAT", year=2026, month=5
    )
    await system_config_repository.set_workday_calendar(
        db_session,
        year=2026,
        entries=[{"date": "20260509", "week": "六", "isHoliday": True, "description": ""}],
        updated_by="test",
    )

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 9),
        end_date=datetime.date(2026, 5, 9),
        format="csv",
    )
    rows = list(csv.reader(io.StringIO(csv_text)))
    data_row = rows[1]
    assert data_row[3] == "2026-05-09"   # 日期
    assert data_row[4] == "星期六"         # 星期
    assert data_row[8] == "正常"           # 狀態 — rest-day work is NORMAL
    assert data_row[9] == "休息日加班"      # 備註
    assert data_row[10] == "3"            # 加班時數


async def test_csv_export_default_excludes_unsubmitted(db_session):
    await _seed_employee_with_shift(
        db_session, emp_id="E042", shift_start_time="09:00", shift_end_time="18:00"
    )
    await summary_repository.upsert_summary(
        db_session,
        emp_id="E042",
        date=datetime.date(2026, 5, 14),
        first_clock_in=datetime.datetime(2026, 5, 14, 9, 0),
        last_clock_out=datetime.datetime(2026, 5, 14, 18, 0),
        status=AttendanceStatus.NORMAL,
    )
    # NO submission for E042

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 14),
        end_date=datetime.date(2026, 5, 14),
        format="csv",
    )
    assert "E042" not in csv_text


async def test_json_export_keeps_english(db_session):
    """Export goes through generate_all_summaries, which recomputes status
    from attendance_logs — so seed real logs (not just a summary row) for
    NORMAL status to survive recomputation.
    """
    from app.models.attendance_log import AttendanceLog, WorkMode

    await _seed_employee_with_shift(
        db_session, emp_id="E043", shift_start_time="09:00", shift_end_time="18:00"
    )
    db_session.add_all(
        [
            AttendanceLog(
                emp_id="E043",
                timestamp=datetime.datetime(2026, 5, 14, 9, 0),
                latitude=25.0,
                longitude=121.5,
                accuracy=10.0,
                ip_address="127.0.0.1",
                work_mode=WorkMode.OFFICE,
            ),
            AttendanceLog(
                emp_id="E043",
                timestamp=datetime.datetime(2026, 5, 14, 18, 0),
                latitude=25.0,
                longitude=121.5,
                accuracy=10.0,
                ip_address="127.0.0.1",
                work_mode=WorkMode.OFFICE,
            ),
        ]
    )
    await db_session.commit()
    await monthly_submission_repository.upsert(
        db_session, emp_id="E043", year=2026, month=5
    )

    json_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 14),
        end_date=datetime.date(2026, 5, 14),
        format="json",
    )
    assert '"emp_id"' in json_text
    assert '"status": "NORMAL"' in json_text
    assert '"shift_time"' in json_text
    assert '"remark"' in json_text
    assert '"submission_status"' in json_text


async def test_export_excludes_absent_on_holiday(db_session):
    """Regression: 5/1 Labour Day must not appear as ABSENT in the export,
    even when an ABSENT row was persisted earlier (e.g. by an admin
    refresh-summaries call when the 2026 calendar wasn't cached yet).
    The export must consult the canonical generate_all_summaries path
    so the holiday guard fires.
    """
    from app.repositories import system_config_repository

    await _seed_employee_with_shift(
        db_session, emp_id="E_HOLIDAY", shift_start_time="09:00", shift_end_time="18:00"
    )
    holiday = datetime.date(2026, 5, 1)
    # Simulate the bug state: an ABSENT row got persisted before calendar load.
    await summary_repository.upsert_summary(
        db_session,
        emp_id="E_HOLIDAY",
        date=holiday,
        first_clock_in=None,
        last_clock_out=None,
        status=AttendanceStatus.ABSENT,
    )
    await monthly_submission_repository.upsert(
        db_session, emp_id="E_HOLIDAY", year=2026, month=5
    )
    # Now seed the calendar with 5/1 as a holiday (what a refresh would produce).
    await system_config_repository.set_workday_calendar(
        db_session,
        year=2026,
        entries=[
            {
                "date": "20260501",
                "week": "五",
                "isHoliday": True,
                "description": "勞動節",
            }
        ],
        updated_by="test",
    )

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=holiday,
        end_date=holiday,
        format="csv",
    )
    # The day appears as a holiday continuity row, NOT as ABSENT/缺勤.
    # The stale ABSENT summary that was persisted earlier is suppressed.
    assert "缺勤" not in csv_text
    assert "E_HOLIDAY" in csv_text
    assert "假日" in csv_text
    assert "勞動節" in csv_text


async def test_export_includes_weekend_continuity_rows(db_session):
    """Weekends appear as 週末 filler rows for every employee in scope so
    payroll can scan a whole month without gaps.
    """
    from app.repositories import system_config_repository

    await _seed_employee_with_shift(
        db_session, emp_id="E_WKND", shift_start_time="09:00", shift_end_time="18:00"
    )
    # Seed a real workday summary on Friday 5/15 (with logs so it survives
    # the recompute) — gives the employee a presence in May so the filler
    # logic emits weekend rows for them.
    from app.models.attendance_log import AttendanceLog, WorkMode
    db_session.add_all(
        [
            AttendanceLog(
                emp_id="E_WKND",
                timestamp=datetime.datetime(2026, 5, 15, 9, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
            AttendanceLog(
                emp_id="E_WKND",
                timestamp=datetime.datetime(2026, 5, 15, 18, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
        ]
    )
    await db_session.commit()
    await monthly_submission_repository.upsert(
        db_session, emp_id="E_WKND", year=2026, month=5
    )
    # Cache the calendar so the loader doesn't hit the network.
    await system_config_repository.set_workday_calendar(
        db_session,
        year=2026,
        entries=[
            {"date": "20260515", "week": "五", "isHoliday": False, "description": ""},
            {"date": "20260516", "week": "六", "isHoliday": True, "description": ""},
            {"date": "20260517", "week": "日", "isHoliday": True, "description": ""},
        ],
        updated_by="test",
    )

    csv_text = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 15),
        end_date=datetime.date(2026, 5, 17),
        format="csv",
    )
    # Friday is a real workday row.
    assert "2026-05-15" in csv_text
    # Sat = 休息日, Sun = 例假日 — new labor-law-correct labels.
    assert "2026-05-16" in csv_text
    assert "2026-05-17" in csv_text
    assert "休息日" in csv_text
    assert "例假日" in csv_text


async def test_xlsx_export_grays_holiday_rows(db_session):
    """The xlsx output applies a gray PatternFill to HOLIDAY/WEEKEND rows
    so reviewers can visually distinguish non-workdays.
    """
    from io import BytesIO
    from openpyxl import load_workbook
    from app.repositories import system_config_repository

    await _seed_employee_with_shift(
        db_session, emp_id="E_XLSX_HOL", shift_start_time="09:00", shift_end_time="18:00"
    )
    # Seed a real workday summary so the employee is in scope for May.
    from app.models.attendance_log import AttendanceLog, WorkMode
    db_session.add_all(
        [
            AttendanceLog(
                emp_id="E_XLSX_HOL",
                timestamp=datetime.datetime(2026, 5, 4, 9, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
            AttendanceLog(
                emp_id="E_XLSX_HOL",
                timestamp=datetime.datetime(2026, 5, 4, 18, 0),
                latitude=25.0, longitude=121.5, accuracy=10.0,
                ip_address="127.0.0.1", work_mode=WorkMode.OFFICE,
            ),
        ]
    )
    await db_session.commit()
    await monthly_submission_repository.upsert(
        db_session, emp_id="E_XLSX_HOL", year=2026, month=5
    )
    await system_config_repository.set_workday_calendar(
        db_session,
        year=2026,
        entries=[
            {"date": "20260501", "week": "五", "isHoliday": True, "description": "勞動節"},
            {"date": "20260504", "week": "一", "isHoliday": False, "description": ""},
        ],
        updated_by="test",
    )

    xlsx_bytes = await reporting_service.export_attendance(
        db_session,
        start_date=datetime.date(2026, 5, 1),
        end_date=datetime.date(2026, 5, 4),
        format="xlsx",
    )
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active

    # Find the 2026-05-01 (holiday) row and the 2026-05-04 (workday) row.
    holiday_row_idx = None
    workday_row_idx = None
    for row_idx in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=4).value
        if date_cell == "2026-05-01":
            holiday_row_idx = row_idx
        elif date_cell == "2026-05-04":
            workday_row_idx = row_idx
    assert holiday_row_idx is not None
    assert workday_row_idx is not None

    holiday_fill = ws.cell(row=holiday_row_idx, column=1).fill
    workday_fill = ws.cell(row=workday_row_idx, column=1).fill
    # Holiday row carries a solid fill; workday row does not.
    assert holiday_fill.fill_type == "solid"
    assert workday_fill.fill_type in (None, "none")
