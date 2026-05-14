"""Reporting service — First-In-Last-Out attendance logic, summary generation, and export.

Provides:
- ``calculate_status``: pure function that classifies a day as NORMAL / LATE /
  EARLY_LEAVE / ABNORMAL based on shift times and clock-in/out timestamps.
- ``generate_daily_summary``: builds (or upserts) a single employee's daily summary.
- ``generate_all_summaries``: batch-generates summaries for every employee on a date.
- ``export_attendance``: exports summary data as CSV or JSON, with optional filters.
"""

import csv
import datetime
import io
import json
from datetime import timedelta
from typing import Literal

from sqlalchemy.ext.asyncio import AsyncSession

from app.models.daily_attendance_summary import AttendanceStatus, DailyAttendanceSummary
from app.repositories import attendance_repository, employee_repository, summary_repository
from app.repositories import (
    monthly_submission_repository,
    reason_repository,
    system_config_repository,
)
from app.utils.taiwan_calendar import (
    DayInfo,
    fetch_calendar_from_cdn,
    is_workday_from_data,
    parse_calendar_json,
)

DEFAULT_GRACE_MINUTES = 5

CHINESE_HEADERS = [
    "員工編號", "姓名", "部門", "日期",
    "班別時間", "上班時間", "下班時間",
    "狀態", "備註", "遲到理由", "送單狀態",
]

STATUS_ZH = {
    "NORMAL": "正常",
    "LATE": "遲到",
    "EARLY_LEAVE": "早退",
    "LATE_AND_EARLY_LEAVE": "遲到且早退",
    "ABNORMAL": "異常",
    "ABSENT": "缺勤",
    "LEAVE": "請假",
    # Synthetic statuses produced only by the export pipeline for non-workday
    # continuity rows — never persisted to daily_attendance_summaries.
    "HOLIDAY": "假日",
    "WEEKEND": "週末",
}

# Statuses that mark a row as a non-working calendar day filler. Used for
# Excel gray-fill styling and to identify filler rows during sort.
_FILLER_STATUSES = frozenset({"HOLIDAY", "WEEKEND"})


def _format_shift_time(start: datetime.time, end: datetime.time) -> str:
    return f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"


def _format_remark(leave_type: str | None, remark: str | None) -> str:
    if leave_type and remark:
        return f"{leave_type}·{remark}"
    return leave_type or remark or ""


def calculate_status(
    shift_start: datetime.time,
    shift_end: datetime.time,
    first_clock_in: datetime.datetime | None,
    last_clock_out: datetime.datetime | None,
    grace_minutes: int = DEFAULT_GRACE_MINUTES,
    leave_type: str | None = None,
) -> AttendanceStatus | None:
    """Determine the attendance status for a single day.

    Parameters
    ----------
    shift_start:
        The employee's scheduled start time.
    shift_end:
        The employee's scheduled end time.
    first_clock_in:
        Earliest punch of the day, or ``None`` if no punches.
    last_clock_out:
        Latest punch of the day, or ``None`` if no punches.
    leave_type:
        Optional non-empty leave-type string (e.g. ``"特休"``). When provided
        and truthy, the function short-circuits and returns
        ``AttendanceStatus.LEAVE`` regardless of punch timing. Empty string
        and ``None`` are treated as "no leave" and fall through to the
        regular timing-based logic.

    Returns
    -------
    AttendanceStatus | None
        ``LEAVE`` when ``leave_type`` is a non-empty string.
        ``None`` when there are no punches at all.
        ``ABNORMAL`` when only one of clock-in / clock-out exists.
        ``LATE`` when clock-in exceeds the grace period (takes precedence).
        ``EARLY_LEAVE`` when clock-out is before shift end.
        ``NORMAL`` otherwise.
    """
    if leave_type:  # truthy → non-empty leave type
        return AttendanceStatus.LEAVE

    if first_clock_in is None and last_clock_out is None:
        return None

    if first_clock_in is None or last_clock_out is None:
        return AttendanceStatus.ABNORMAL

    # Calculate grace deadline: shift_start + grace period
    # Use a reference date to combine time + timedelta
    reference_date = first_clock_in.date()
    grace_deadline = (
        datetime.datetime.combine(reference_date, shift_start)
        + timedelta(minutes=grace_minutes)
    ).time()

    is_late = first_clock_in.time() > grace_deadline

    # Only one punch exists — still check for lateness so the employee
    # can submit a reason immediately after a late clock-in.
    # A single on-time punch is a normal clock-in (not yet clocked out).
    if first_clock_in == last_clock_out:
        return AttendanceStatus.LATE if is_late else AttendanceStatus.NORMAL

    is_early_leave = last_clock_out.time() < shift_end

    if is_late and is_early_leave:
        return AttendanceStatus.LATE_AND_EARLY_LEAVE

    if is_late:
        return AttendanceStatus.LATE

    if is_early_leave:
        return AttendanceStatus.EARLY_LEAVE

    return AttendanceStatus.NORMAL


async def generate_daily_summary(
    session: AsyncSession,
    emp_id: str,
    date: datetime.date,
) -> DailyAttendanceSummary | None:
    """Build or upsert a daily attendance summary for one employee.

    Returns ``None`` if the employee has no attendance logs AND no existing
    leave_type on *date*. Preserves any pre-set leave_type/remark across recompute.
    """
    employee = await employee_repository.find_by_id(session, emp_id)
    if employee is None:
        return None

    # Read any existing summary's leave_type/remark so we preserve them
    existing_rows = await summary_repository.find_by_employee(
        session, emp_id, start_date=date, end_date=date
    )
    existing_leave_type = existing_rows[0].leave_type if existing_rows else None
    existing_remark = existing_rows[0].remark if existing_rows else None

    first_log = await attendance_repository.find_first_clock_in(session, emp_id, date)
    last_log = await attendance_repository.find_last_clock_out(session, emp_id, date)

    first_clock_in = first_log.timestamp if first_log is not None else None
    last_clock_out = last_log.timestamp if last_log is not None else None

    grace_minutes = await system_config_repository.get_grace_period(session)
    status = calculate_status(
        employee.shift_start_time,
        employee.shift_end_time,
        first_clock_in,
        last_clock_out,
        grace_minutes=grace_minutes,
        leave_type=existing_leave_type,
    )

    if status is None:
        return None

    summary = await summary_repository.upsert_summary(
        session,
        emp_id=emp_id,
        date=date,
        first_clock_in=first_clock_in,
        last_clock_out=last_clock_out,
        status=status,
        leave_type=existing_leave_type,
        remark=existing_remark,
    )

    return summary


async def _load_calendar_for_year(
    session: AsyncSession, year: int
) -> list[DayInfo]:
    """Load Taiwan workday calendar for a year, auto-fetching from the CDN
    on cache miss and persisting the result.

    Without this auto-fetch, an empty cache silently degrades to a Mon-Fri
    fallback in ``is_workday_from_data``, which incorrectly classifies
    national holidays (e.g. 5/1 Labour Day on a Friday) as workdays and
    generates spurious ABSENT summaries. Mirrors the behavior of
    ``GET /api/config/workdays`` so reports/exports don't depend on
    someone having visited the monthly-override page first.

    Returns an empty list only if both the cache is empty AND the CDN
    fetch fails — in that edge case the Mon-Fri fallback applies.
    """
    cached = await system_config_repository.get_workday_calendar(session, year)
    if cached is not None:
        raw_entries = cached.get("entries", []) if isinstance(cached, dict) else []
        return parse_calendar_json(raw_entries)

    # Cache miss — fetch from CDN and persist.
    data = await fetch_calendar_from_cdn(year)
    if not data:
        return []

    raw_entries = [
        {
            "date": d.date.strftime("%Y%m%d"),
            "week": d.weekday_zh,
            "isHoliday": d.is_holiday,
            "description": d.description,
        }
        for d in data
    ]
    await system_config_repository.set_workday_calendar(
        session, year, raw_entries, updated_by="reporting_service"
    )
    return data


async def generate_all_summaries(
    session: AsyncSession,
    date: datetime.date,
) -> list[DailyAttendanceSummary]:
    """Generate daily summaries for every employee on *date*.

    Phase 12 — ABSENT tracking:
    * Employees who punched get a NORMAL/LATE/etc. summary (unchanged).
    * On workdays, employees who did NOT punch get an ABSENT summary.
    * On holidays / weekends (per Taiwan calendar, fallback Mon-Fri),
      non-punching employees are silently skipped — no ABSENT generation.

    Phase 14E — Terminated employees:
    * Terminated employees' existing punched summaries are still returned
      (required by LSA §30(5) retention).
    * ABSENT is NOT generated for employees terminated on or before *date*
      (they weren't active — can't be absent).
    """
    # Include terminated so historical summaries are returned. ABSENT generation
    # below is guarded separately.
    employees = await employee_repository.find_all(
        session, skip=0, limit=10000, include_terminated=True
    )

    summaries: list[DailyAttendanceSummary] = []
    handled_emp_ids: set[str] = set()  # punched OR on leave

    for emp in employees:
        summary = await generate_daily_summary(session, emp.emp_id, date)
        if summary is not None:
            summaries.append(summary)
            handled_emp_ids.add(emp.emp_id)

    # Determine workday status via cached Taiwan calendar (fallback Mon-Fri)
    calendar_data = await _load_calendar_for_year(session, date.year)
    if not is_workday_from_data(calendar_data, date):
        return summaries

    # Generate ABSENT summaries for employees who did not punch
    for emp in employees:
        if emp.emp_id in handled_emp_ids:
            continue
        # Skip ABSENT generation for employees already terminated on this date
        if emp.terminated_at is not None and emp.terminated_at.date() <= date:
            continue
        absent_summary = await summary_repository.upsert_summary(
            session,
            emp_id=emp.emp_id,
            date=date,
            first_clock_in=None,
            last_clock_out=None,
            status=AttendanceStatus.ABSENT,
        )
        summaries.append(absent_summary)

    return summaries


async def get_daily_report(
    session: AsyncSession,
    start_date: datetime.date,
    end_date: datetime.date | None = None,
    department: str | None = None,
    emp_id: str | None = None,
    status_filter: str | None = None,
    include_terminated: bool = False,
    submission_filter: Literal["submitted", "unsubmitted", "all"] = "submitted",
) -> list[DailyAttendanceSummary]:
    """Return daily attendance summaries for a date range (inclusive).

    Parameters
    ----------
    session:
        The async database session.
    start_date:
        The first date to retrieve summaries for.
    end_date:
        The last date (inclusive). Defaults to *start_date* if ``None``.
    department:
        Optional department filter.
    emp_id:
        Optional individual employee filter. When provided, terminated
        employees are always included (LSA §30(5) retention).
    status_filter:
        Optional status filter (e.g. ``"LATE"``).
    include_terminated:
        When ``False`` (default) and no explicit ``emp_id`` is provided,
        terminated employees are excluded from the result. Ignored when
        ``emp_id`` is set.
    submission_filter:
        Controls visibility based on the per-month submission state in
        ``monthly_submissions``. One of:

        * ``"submitted"`` (default) — only summaries whose
          (emp_id, year, month) has a submission row.
        * ``"unsubmitted"`` — only summaries with no submission row.
        * ``"all"`` — no submission-state filtering.

    Returns
    -------
    list[DailyAttendanceSummary]
        Matching summaries, sorted by date then emp_id.
    """
    if end_date is None:
        end_date = start_date

    all_summaries: list[DailyAttendanceSummary] = []
    current = start_date
    while current <= end_date:
        day_summaries = await generate_all_summaries(session, current)
        all_summaries.extend(day_summaries)
        current += datetime.timedelta(days=1)

    if emp_id is not None:
        # Explicit emp_id filter always honored (terminated or not — LSA retention).
        all_summaries = [s for s in all_summaries if s.emp_id == emp_id]
    elif not include_terminated:
        # No specific emp_id picked: hide terminated employees by default.
        terminated_emp_ids = await employee_repository.find_terminated_ids(session)
        all_summaries = [s for s in all_summaries if s.emp_id not in terminated_emp_ids]

    if department is not None:
        employees = await employee_repository.find_by_department(
            session, department, include_terminated=include_terminated
        )
        dept_emp_ids = {emp.emp_id for emp in employees}
        all_summaries = [s for s in all_summaries if s.emp_id in dept_emp_ids]

    if status_filter is not None:
        try:
            target_status = AttendanceStatus(status_filter)
        except ValueError:
            return []
        all_summaries = [s for s in all_summaries if s.status == target_status]

    if submission_filter != "all":
        cache: dict[tuple[int, int], set[str]] = {}

        async def _is_submitted(emp: str, d: datetime.date) -> bool:
            key = (d.year, d.month)
            if key not in cache:
                cache[key] = await monthly_submission_repository.submitted_emp_ids(
                    session, year=d.year, month=d.month
                )
            return emp in cache[key]

        filtered: list[DailyAttendanceSummary] = []
        for s in all_summaries:
            submitted = await _is_submitted(s.emp_id, s.date)
            if submission_filter == "submitted" and submitted:
                filtered.append(s)
            elif submission_filter == "unsubmitted" and not submitted:
                filtered.append(s)
        all_summaries = filtered

    all_summaries.sort(key=lambda s: (s.date, s.emp_id))
    return all_summaries


async def export_attendance(
    session: AsyncSession,
    start_date: datetime.date,
    end_date: datetime.date,
    format: str,
    department: str | None = None,
    emp_id: str | None = None,
    include_terminated: bool = False,
    submission_filter: Literal["submitted", "unsubmitted", "all"] = "submitted",
) -> str | bytes:
    """Export attendance summaries as CSV, JSON, or Excel.

    CSV / Excel output uses Chinese headers and Chinese-localized status and
    submission values. JSON output preserves English keys and raw enum values.

    Columns (CSV/Excel):
        員工編號, 姓名, 部門, 日期, 班別時間, 上班時間, 下班時間,
        狀態, 備註, 遲到理由, 送單狀態

    Parameters
    ----------
    session:
        The async database session.
    start_date / end_date:
        Inclusive date range.
    format:
        ``"csv"``, ``"json"``, or ``"xlsx"``.
    department:
        Optional department filter.
    emp_id:
        Optional individual employee filter (always honored — LSA retention).
    include_terminated:
        When ``False`` (default) and no explicit ``emp_id`` is provided,
        terminated employees are excluded.
    submission_filter:
        Controls visibility based on per-month submission state. One of
        ``"submitted"`` (default), ``"unsubmitted"``, or ``"all"``.

    Returns
    -------
    str | bytes
        Formatted string (CSV/JSON) or bytes (xlsx).
    """
    # --- gather summaries via get_daily_report so the calendar / holiday /
    #     submission_filter logic stays in one place. Reading
    #     daily_attendance_summaries directly would leak stale ABSENT rows
    #     persisted before the Taiwan calendar was cached (e.g. 5/1 Labour
    #     Day was generated as ABSENT under the Mon-Fri fallback).
    all_summaries = await get_daily_report(
        session,
        start_date=start_date,
        end_date=end_date,
        department=department,
        emp_id=emp_id,
        include_terminated=include_terminated,
        submission_filter=submission_filter,
    )

    # --- determine employees in scope for the *export* (independent of
    #     whether each one happens to have a real summary in this range).
    #     Needed so the holiday/weekend filler logic can emit continuity
    #     rows even when the entire range falls on non-workdays.
    if emp_id is not None:
        scope_emp = await employee_repository.find_by_id(session, emp_id)
        scope_employees = [scope_emp] if scope_emp is not None else []
    elif department is not None:
        scope_employees = await employee_repository.find_by_department(
            session, department, include_terminated=include_terminated
        )
    else:
        scope_employees = await employee_repository.find_all(
            session, skip=0, limit=10000, include_terminated=include_terminated
        )
    emp_map: dict[str, object] = {e.emp_id: e for e in scope_employees}

    # Cover anyone present in real summaries who somehow slipped through
    # the scope query (e.g. an emp_id explicit filter where the row is
    # already there).
    for eid in {s.emp_id for s in all_summaries}:
        if eid not in emp_map:
            emp = await employee_repository.find_by_id(session, eid)
            if emp is not None:
                emp_map[eid] = emp

    # --- preload reasons keyed by summary_id ---
    summary_ids = [s.id for s in all_summaries if s.id is not None]
    reasons = await reason_repository.find_by_summary_ids(session, summary_ids)
    reason_map: dict[int, str] = {r.summary_id: r.reason for r in reasons}

    # --- per-(year, month) submission cache (used only for the row-level
    #     submission_status label; the list is already filtered by
    #     submission_filter inside get_daily_report).
    sub_cache: dict[tuple[int, int], set[str]] = {}

    async def _is_submitted(e: str, d: datetime.date) -> bool:
        key = (d.year, d.month)
        if key not in sub_cache:
            sub_cache[key] = await monthly_submission_repository.submitted_emp_ids(
                session, year=d.year, month=d.month
            )
        return e in sub_cache[key]

    # --- build English-keyed rows (JSON path uses these directly) ---
    rows: list[dict[str, str]] = []
    for s in all_summaries:
        emp = emp_map.get(s.emp_id)
        sub = await _is_submitted(s.emp_id, s.date)
        rows.append({
            "emp_id": s.emp_id,
            "name": emp.name if emp else "",
            "department": emp.department if emp else "",
            "date": s.date.isoformat(),
            "shift_time": (
                _format_shift_time(emp.shift_start_time, emp.shift_end_time)
                if emp
                else ""
            ),
            "first_clock_in": (
                s.first_clock_in.isoformat() if s.first_clock_in else ""
            ),
            "last_clock_out": (
                s.last_clock_out.isoformat() if s.last_clock_out else ""
            ),
            "status": s.status.value,
            "remark": _format_remark(s.leave_type, s.remark),
            "reason": reason_map.get(s.id or -1, ""),
            "submission_status": "submitted" if sub else "unsubmitted",
        })

    # --- synthesize holiday / weekend continuity rows ---
    # The on-screen Reports page deliberately skips these; only the exported
    # file gets them, so reviewers can scan a whole month without mentally
    # filling in the gaps. Filler rows are never persisted.
    #
    # Classification:
    #   * Day in calendar with is_holiday=True and a description → "HOLIDAY"
    #     (e.g. 勞動節, 端午節) — description goes into the remark column.
    #   * Day in calendar with is_holiday=True and empty description → "WEEKEND"
    #     (calendar entry exists but it's just a regular Sat/Sun).
    #   * Day not in calendar at all → fall back to weekday() >= 5 = WEEKEND.
    #   * 補班 (is_makeup_workday=True) → workday, no filler row needed.
    #
    # Filler rows respect the same employee scope as the real rows (emp_map)
    # and the same submission_filter (we skip filler rows for an employee
    # whose month doesn't satisfy the filter).
    if emp_map:
        years_in_range: set[int] = set()
        cursor = start_date
        while cursor <= end_date:
            years_in_range.add(cursor.year)
            cursor += timedelta(days=1)
        calendars: dict[int, list[DayInfo]] = {
            y: await _load_calendar_for_year(session, y) for y in years_in_range
        }

        existing_pairs: set[tuple[str, datetime.date]] = {
            (r["emp_id"], datetime.date.fromisoformat(r["date"])) for r in rows
        }

        cursor = start_date
        while cursor <= end_date:
            cal = calendars.get(cursor.year, [])
            day_info = next((di for di in cal if di.date == cursor), None)

            # 補班 (make-up workday) is a workday — skip filler.
            if day_info is not None and day_info.is_makeup_workday:
                cursor += timedelta(days=1)
                continue

            # Determine workday status.
            if day_info is not None:
                non_workday = day_info.is_holiday
            else:
                non_workday = cursor.weekday() >= 5

            if not non_workday:
                cursor += timedelta(days=1)
                continue

            # Classify HOLIDAY vs WEEKEND.
            if day_info is not None and day_info.description:
                filler_status = "HOLIDAY"
                filler_remark = day_info.description
            else:
                filler_status = "WEEKEND"
                filler_remark = ""

            for emp_id_iter, emp in emp_map.items():
                if (emp_id_iter, cursor) in existing_pairs:
                    continue
                sub = await _is_submitted(emp_id_iter, cursor)
                # Honor submission_filter: skip filler rows that the user
                # has explicitly filtered out.
                if submission_filter == "submitted" and not sub:
                    continue
                if submission_filter == "unsubmitted" and sub:
                    continue
                rows.append({
                    "emp_id": emp_id_iter,
                    "name": emp.name,
                    "department": emp.department,
                    "date": cursor.isoformat(),
                    "shift_time": _format_shift_time(
                        emp.shift_start_time, emp.shift_end_time
                    ),
                    "first_clock_in": "",
                    "last_clock_out": "",
                    "status": filler_status,
                    "remark": filler_remark,
                    "reason": "",
                    "submission_status": "submitted" if sub else "unsubmitted",
                })

            cursor += timedelta(days=1)

    # Re-sort so filler rows interleave naturally with real rows by date.
    rows.sort(key=lambda r: (r["date"], r["emp_id"]))

    if format == "json":
        return json.dumps(rows, ensure_ascii=False)

    # --- translate to Chinese for CSV / Excel ---
    zh_rows: list[list[str]] = []
    for row in rows:
        zh_rows.append([
            row["emp_id"],
            row["name"],
            row["department"],
            row["date"],
            row["shift_time"],
            row["first_clock_in"],
            row["last_clock_out"],
            STATUS_ZH.get(row["status"], row["status"]),
            row["remark"],
            row["reason"],
            "已送單" if row["submission_status"] == "submitted" else "未送單",
        ])

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        bold_font = Font(bold=True)
        for col_idx, header in enumerate(CHINESE_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        # Light gray fill for holiday / weekend filler rows, so reviewers can
        # visually distinguish "this day is not a workday" from real
        # attendance rows at a glance.
        filler_fill = PatternFill(
            fill_type="solid", start_color="FFEFEFEF", end_color="FFEFEFEF"
        )

        # We rely on rows[] (the English-keyed dicts) and zh_rows[] (the
        # Chinese-localized lists) staying index-aligned — they are produced
        # together by the same loop above.
        for row_idx, (row_dict, row_values) in enumerate(zip(rows, zh_rows), 2):
            is_filler = row_dict["status"] in _FILLER_STATUSES
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if is_filler:
                    cell.fill = filler_fill

        for col_idx, header in enumerate(CHINESE_HEADERS, 1):
            max_len = len(header)
            for row_values in zh_rows:
                max_len = max(max_len, len(str(row_values[col_idx - 1])))
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = max_len + 2

        ws.auto_filter.ref = ws.dimensions

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CHINESE_HEADERS)
    for row_values in zh_rows:
        writer.writerow(row_values)

    return output.getvalue()
